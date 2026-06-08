import scipy.io
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

mat = scipy.io.loadmat('WDF_cornering_sweep.mat', simplify_cells=True)
all_data = mat['allCorneringData']
wdf_values = mat['WDF'].flatten().astype(int)

COLS = ['Radius', 'steering', 'speed', 'latG', 'USG', 'Beta', 'af', 'ar', 'WDF']
COL_LABELS = {
    'Radius':   'Radius (ft)',
    'steering': 'Steering Angle (deg)',
    'speed':    'Speed (ft/s)',
    'latG':     'Lateral G',
    'USG':      'Understeer Gradient',
    'Beta':     'Beta - Chassis Slip (deg)',
    'af':       'Front Slip Angle (deg)',
    'ar':       'Rear Slip Angle (deg)',
    'WDF':      'WDF (%)',
}

# Styles
HDR_FILL   = PatternFill('solid', start_color='1F3864')   # dark navy
HDR_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
DATA_FONT  = Font(name='Arial', size=10)
ALT_FILL   = PatternFill('solid', start_color='DCE6F1')   # light blue
thin = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center')

NUM_FMT = {
    'Radius':   '0',
    'steering': '0.000',
    'speed':    '0.00',
    'latG':     '0.0000',
    'USG':      '0.0000',
    'Beta':     '0.000',
    'af':       '0.000',
    'ar':       '0.000',
    'WDF':      '0',
}

COL_WIDTHS = {
    'Radius': 12, 'steering': 22, 'speed': 14, 'latG': 13,
    'USG': 22, 'Beta': 24, 'af': 22, 'ar': 22, 'WDF': 10,
}

wb = Workbook()
wb.remove(wb.active)   # remove default blank sheet

for wdf in wdf_values:
    field = f'WDF_{wdf}'
    tbl = all_data[field]
    sheet_name = f'WDF = {wdf}%'
    ws = wb.create_sheet(title=sheet_name)

    # --- Header row ---
    for col_idx, col in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=COL_LABELS[col])
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col]
    ws.row_dimensions[1].height = 20

    # --- Data rows ---
    # tbl values may be a dict of 1-D arrays
    n_rows = len(np.atleast_1d(tbl[COLS[0]]))
    for row_idx in range(n_rows):
        excel_row = row_idx + 2
        fill = ALT_FILL if row_idx % 2 == 1 else None
        for col_idx, col in enumerate(COLS, start=1):
            val = float(np.atleast_1d(tbl[col])[row_idx])
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.number_format = NUM_FMT[col]
            cell.alignment = center
            cell.border = border
            if fill:
                cell.fill = fill

    # Freeze the header row
    ws.freeze_panes = 'A2'

wb.save('WDF_Cornering_Sweep.xlsx')
print(f'Saved WDF_Cornering_Sweep.xlsx with {len(wdf_values)} sheets')
